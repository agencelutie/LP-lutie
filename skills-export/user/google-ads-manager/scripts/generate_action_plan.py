#!/usr/bin/env python3
"""
Génère le plan d'action Excel à partir du JSON d'audit.
Usage: python generate_action_plan.py --audit-json <path> --output <path.xlsx>

Ce script peut aussi être utilisé avec des données passées directement via stdin JSON.
"""
import argparse
import json
import sys
import os
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ pip install openpyxl --break-system-packages")
    sys.exit(1)


# ─── Couleurs & styles ───────────────────────────────────────────────────────

COLOR_P1 = "FFCCCC"       # Rouge clair
COLOR_P2 = "FFF2CC"       # Jaune clair
COLOR_P3 = "D9EAD3"       # Vert clair
COLOR_HEADER = "1F4E79"   # Bleu foncé
COLOR_HEADER_FONT = "FFFFFF"
COLOR_SECTION = "BDD7EE"  # Bleu clair pour les sous-titres de section

FONT_HEADER = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=11)
FONT_SECTION = Font(name="Calibri", bold=True, color="1F4E79", size=10)
FONT_BODY = Font(name="Calibri", size=10)
FONT_BODY_BOLD = Font(name="Calibri", bold=True, size=10)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_header_style(cell, text):
    cell.value = text
    cell.font = FONT_HEADER
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER


def apply_cell_style(cell, text, priority=None, bold=False):
    cell.value = text
    cell.font = FONT_BODY_BOLD if bold else FONT_BODY
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    if priority == "P1":
        cell.fill = PatternFill("solid", fgColor=COLOR_P1)
    elif priority == "P2":
        cell.fill = PatternFill("solid", fgColor=COLOR_P2)
    elif priority == "P3":
        cell.fill = PatternFill("solid", fgColor=COLOR_P3)


def priority_label(p):
    labels = {"P1": "🔴 P1 — Urgent", "P2": "🟡 P2 — Ce mois", "P3": "🟢 P3 — Continu"}
    return labels.get(p, p)


def generate_excel(actions: list, summary: dict, output_path: str, account_name: str = ""):
    """Génère le fichier Excel du plan d'action."""
    wb = openpyxl.Workbook()

    # ─── Feuille 1 : Plan d'action ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Plan d'Action"

    # Titre
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Plan d'Action Google Ads{' — ' + account_name if account_name else ''} | {datetime.now().strftime('%d/%m/%Y')}"
    title_cell.font = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=14)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = ALIGN_CENTER

    ws.row_dimensions[1].height = 30

    # En-têtes colonnes
    headers = ["#", "Priorité", "Catégorie", "Action", "Détail / Justification",
               "Impact Estimé", "Effort", "Responsable", "Deadline", "Statut"]
    col_widths = [5, 18, 16, 35, 50, 28, 10, 16, 14, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_idx)
        apply_header_style(cell, header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 30

    # Trier : P1 d'abord, puis P2, puis P3
    priority_order = {"P1": 0, "P2": 1, "P3": 2}
    sorted_actions = sorted(actions, key=lambda x: priority_order.get(x.get("priority", "P3"), 3))

    current_row = 3
    for idx, action in enumerate(sorted_actions, start=1):
        p = action.get("priority", "P3")
        row_data = [
            idx,
            priority_label(p),
            action.get("category", ""),
            action.get("action", ""),
            action.get("detail", ""),
            action.get("impact", ""),
            action.get("effort", ""),
            action.get("owner", ""),
            action.get("deadline", ""),
            action.get("status", "À faire"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            apply_cell_style(cell, value, priority=p)
        ws.row_dimensions[current_row].height = 40
        current_row += 1

    ws.freeze_panes = "A3"

    # ─── Feuille 2 : Résumé de l'audit ─────────────────────────────────────
    ws2 = wb.create_sheet("Résumé Audit")

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Résumé de l'Audit Google Ads"
    ws2["A1"].font = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=13)
    ws2["A1"].fill = PatternFill("solid", fgColor=COLOR_HEADER)
    ws2["A1"].alignment = ALIGN_CENTER

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 40

    summary_rows = [
        ("📅 Période analysée", summary.get("period", "30 derniers jours"), "", ""),
        ("💰 Dépense totale", f"{summary.get('total_spend_eur', 0):,.2f} €", "", ""),
        ("🎯 Conversions", summary.get("total_conversions", "N/A"), "", ""),
        ("📊 CPA moyen", f"{summary.get('avg_cpa_eur', 'N/A'):,.2f} €" if summary.get("avg_cpa_eur") else "N/A", "", ""),
        ("📈 ROAS", summary.get("roas", "N/A"), "", ""),
        ("", "", "", ""),
        ("🔴 Problèmes critiques (P1)", summary.get("p1_count", 0), "", "À traiter cette semaine"),
        ("🟡 Avertissements (P2)", summary.get("p2_count", 0), "", "À traiter ce mois"),
        ("🟢 Améliorations (P3)", summary.get("p3_count", 0), "", "Optimisation continue"),
        ("", "", "", ""),
        ("🏆 Score global", summary.get("score", "—"), "", summary.get("score_comment", "")),
    ]

    for row_idx, (label, value, _, note) in enumerate(summary_rows, start=3):
        ws2.cell(row=row_idx, column=1).value = label
        ws2.cell(row=row_idx, column=1).font = FONT_BODY_BOLD
        ws2.cell(row=row_idx, column=1).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2).value = value
        ws2.cell(row=row_idx, column=2).font = FONT_BODY
        ws2.cell(row=row_idx, column=2).border = THIN_BORDER
        ws2.cell(row=row_idx, column=4).value = note
        ws2.cell(row=row_idx, column=4).font = FONT_BODY
        ws2.cell(row=row_idx, column=4).border = THIN_BORDER

    wb.save(output_path)
    print(f"✅ Plan d'action généré : {output_path}")
    print(f"   {len(sorted_actions)} actions | {sum(1 for a in actions if a.get('priority')=='P1')} P1 | {sum(1 for a in actions if a.get('priority')=='P2')} P2 | {sum(1 for a in actions if a.get('priority')=='P3')} P3")


# ─── Actions de démonstration (si pas de JSON d'audit) ───────────────────────

DEMO_ACTIONS = [
    {
        "priority": "P1", "category": "Tracking",
        "action": "Corriger le tag de conversion dupliqué",
        "detail": "La conversion 'Achat' est comptée 2x (tag direct + import GA4). Supprimer l'une des deux sources.",
        "impact": "Données fiables = optimisation enchères correcte",
        "effort": "Faible", "owner": "Dev + PPC", "deadline": "Cette semaine", "status": "À faire"
    },
    {
        "priority": "P1", "category": "Budget",
        "action": "Réallouer budget de la campagne Display (0 conv.) vers Search Brand",
        "detail": "Campagne Display a dépensé 450€ sur 30j avec 0 conversions. Search Brand est limitée par son budget.",
        "impact": "Récupération ~450€/mois sur un canal performant",
        "effort": "Faible", "owner": "PPC Manager", "deadline": "Cette semaine", "status": "À faire"
    },
    {
        "priority": "P1", "category": "Search Terms",
        "action": "Ajouter 23 négatifs identifiés dans le rapport search terms",
        "detail": "23 requêtes non pertinentes ont généré 280€ de clics sans conversion. Voir liste jointe.",
        "impact": "Économie estimée ~280€/mois",
        "effort": "Faible", "owner": "PPC Manager", "deadline": "Cette semaine", "status": "À faire"
    },
    {
        "priority": "P2", "category": "Annonces",
        "action": "Améliorer la qualité des RSA de 3 groupes d'annonces (statut 'Poor')",
        "detail": "3 groupes ont leur RSA principal en statut 'Poor'. Diversifier les titres et inclure le mot-clé principal.",
        "impact": "Amélioration Quality Score → CPC réduit",
        "effort": "Moyen", "owner": "Copywriter + PPC", "deadline": "Dans 2 semaines", "status": "À faire"
    },
    {
        "priority": "P2", "category": "Extensions",
        "action": "Ajouter les sitelinks manquants sur 4 campagnes",
        "detail": "4 campagnes actives n'ont que 2 sitelinks. Minimum requis : 6. Impact direct sur le taux d'affichage des extensions.",
        "impact": "CTR +5-15% estimé",
        "effort": "Faible", "owner": "PPC Manager", "deadline": "Cette semaine", "status": "À faire"
    },
    {
        "priority": "P3", "category": "Mots-clés",
        "action": "Tester l'expansion Broad Match sur campagne Search principale",
        "detail": "La campagne a 45 conv/mois avec Target CPA. Conditions réunies pour tester le broad match sur 5 adgroups clés.",
        "impact": "Volume +20-40% si ROAS maintenu",
        "effort": "Moyen", "owner": "PPC Manager", "deadline": "Mois prochain", "status": "À planifier"
    },
]

DEMO_SUMMARY = {
    "period": "30 derniers jours",
    "total_spend_eur": 3240.50,
    "total_conversions": 87,
    "avg_cpa_eur": 37.24,
    "roas": "3.2x",
    "p1_count": 3,
    "p2_count": 2,
    "p3_count": 1,
    "score": "62/100",
    "score_comment": "Compte fonctionnel mais avec des pertes évitables (~730€/mois récupérables rapidement)"
}


def main():
    parser = argparse.ArgumentParser(description="Génère le plan d'action Google Ads en Excel")
    parser.add_argument("--audit-json", help="Chemin vers le JSON d'audit (optionnel, utilise des données de démo si absent)")
    parser.add_argument("--output", default="./plan_action_google_ads.xlsx", help="Chemin du fichier Excel de sortie")
    parser.add_argument("--account-name", default="", help="Nom du compte (pour le titre)")
    parser.add_argument("--actions-json", help="JSON des actions à intégrer (passé directement)")
    args = parser.parse_args()

    actions = DEMO_ACTIONS
    summary = DEMO_SUMMARY

    if args.actions_json:
        try:
            data = json.loads(args.actions_json)
            actions = data.get("actions", DEMO_ACTIONS)
            summary = data.get("summary", DEMO_SUMMARY)
        except json.JSONDecodeError as e:
            print(f"❌ Erreur parsing JSON des actions : {e}")
            sys.exit(1)
    elif args.audit_json:
        try:
            with open(args.audit_json, "r", encoding="utf-8") as f:
                audit_data = json.load(f)
            # Si le JSON contient déjà les actions structurées
            if "actions" in audit_data:
                actions = audit_data["actions"]
                summary = audit_data.get("summary", DEMO_SUMMARY)
            else:
                print("ℹ️  JSON d'audit chargé. Utilisez le cadre d'audit pour structurer les actions.")
                print("   Génération avec les données d'audit brutes...")
                summary = audit_data.get("summary", DEMO_SUMMARY)
        except FileNotFoundError:
            print(f"❌ Fichier non trouvé : {args.audit_json}")
            sys.exit(1)
    else:
        print("ℹ️  Aucun JSON d'audit fourni — génération avec des données de démonstration.")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    generate_excel(actions, summary, args.output, args.account_name)


if __name__ == "__main__":
    main()
